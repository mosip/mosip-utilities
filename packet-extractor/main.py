import sys
import os
import argparse
import urllib.parse
import requests
import json
from openpyxl.styles import Alignment
from openpyxl import load_workbook
from datetime import timedelta, datetime
from threading import Thread
import base64
import xmltodict
import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

import config as conf
from utils.app_db_helper import fetch_all
from utils.app_path import log_path, source_path, output_folder_path, get_custom_logpath
from utils.app_logger import init_logger, info, error, debug
from utils.app_file_helper import read_lines
from utils.app_session import MosipSession
from utils.app_helper import time_diff, get_time_in_sec, get_timestamp, parse_response
from utils.app_json import dict_to_json
from utils.app_csv import read_csv_file

result_list = []

class ParentThread(Thread):
  def __init__(self, auth_token, rid, logger_name):
    Thread.__init__(self)
    self.auth_token = auth_token
    self.rid = rid
    self.logger_name = logger_name

  def run(self):
    process_rid(self.rid, self.auth_token, self.logger_name)

class IdRepoThread(Thread):
  def __init__(self, auth_token, rid, logger_name):
    Thread.__init__(self)
    self.auth_token = auth_token
    self.rid = rid
    self.value = None
    self.logger_name = logger_name

  def run(self):
    self.value = get_idrepo_identity_by_rid(self.auth_token, self.rid, self.logger_name)

class PacketManagerThread(Thread):
  def __init__(self, auth_token, rid, logger_name):
    Thread.__init__(self)
    self.auth_token = auth_token
    self.rid = rid
    self.value = None
    self.logger_name = logger_name

  def run(self):
    self.value =get_info_from_packet(self.auth_token, ["firstName",
         "lastName","middleName","gender","presentProvince"], self.rid, "NEW", self.logger_name)

class PacketManagerBiometricsThread(Thread):
  def __init__(self, auth_token, rid, logger_name):
    Thread.__init__(self)
    self.auth_token = auth_token
    self.rid = rid
    self.value = None
    self.logger_name = logger_name

  def run(self):
    self.value =get_biometrics_from_packet(self.auth_token, self.rid, "NEW", self.logger_name)

def args_parse():
    parser = argparse.ArgumentParser()
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument('--file', action='store_true',  help='Process starts from file')
    group.add_argument('--db', action='store_true',  help='Process starts from db')
    # group.add_argument('--auth_uin', action='store_true',  help='Create credential for auth(UIN)')
    args = parser.parse_args()
    return args, parser

def get_auth_token(logger_name):
    start_time = get_time_in_sec()
    try:
        regproc_ms = MosipSession(conf.auth_server_url, conf.regproc_client_id, conf.regproc_secret_key, logger_name, conf.regproc_app_id)
        return regproc_ms.token
    except Exception as e:
        error(logger_name, f"Exception while getting auth token - {e}")

def main():
    args, parser = args_parse()
    try:
        if args.db:
            process_from_db()
        elif args.file:
            process_from_file()
    except Exception as e:
        print("Exception occurred", e)
    finally:
        sys.exit(0)

def process_from_file():
    start_time = get_time_in_sec()
    logger_name = "single"
    try:
        init_logger(logger_name, log_file=log_path, level=conf.logger_level)
        if (source_path):
            dict_rows = read_csv_file(source_path)
            if (len(dict_rows) > 0):
                info(logger_name, f"Number of RIDs to process: {len(dict_rows)}")
                curdate = datetime.today().strftime("%Y%m%d%H%M")
                result_file_name = f"result_{curdate}"
                output_path = os.path.join(output_folder_path, result_file_name + ".xlsx")
                wb = load_workbook("./resource/template.xlsx")
                wb.save(output_path)
                process_rids(dict_rows, output_path, logger_name)                
        else:
            info(logger_name, f"Source file not present")            
    except Exception as e:
        print(e)
        error(logger_name, e)
    finally:
        prev_time, prstr = time_diff(start_time)
        info(logger_name, "Total time taken by the script: " + prstr)

def process_from_db():
    start_time = get_time_in_sec()
    uin_gen_date = conf.start_date
    logger_name = uin_gen_date    
    try:
        end_date = conf.end_date
        if (not end_date):
            end_date = datetime.now().strftime("%Y%m%d")
        while(True):
            print(uin_gen_date, end_date)
            logger_name = uin_gen_date
            init_logger(logger_name, log_file=get_custom_logpath(uin_gen_date), level=conf.logger_level)
            uin_gen_start_date = datetime.strptime(uin_gen_date,"%Y%m%d").strftime("%Y-%m-%d")
            uin_gen_end_date = (datetime.strptime(uin_gen_date,"%Y%m%d") + timedelta(days=1)).strftime("%Y-%m-%d")
            rid_list = fetch_rids(uin_gen_start_date,uin_gen_end_date)
            if (len(rid_list) > 0):
                info(logger_name, f"Number of RIDs to process: {len(rid_list)}")
                result_file_name = f"result_{uin_gen_date}"
                output_path = os.path.join(output_folder_path, uin_gen_date, result_file_name + ".xlsx")
                wb = load_workbook("./resource/template.xlsx")
                wb.save(output_path)
                process_rids(rid_list, output_path, logger_name)
            else:
                info(logger_name, f'No rids found on {uin_gen_date}')
            if (uin_gen_date == end_date):
                break
            
            uin_gen_date = (datetime.strptime(uin_gen_date,"%Y%m%d") + timedelta(days=1)).strftime("%Y%m%d")        
    except Exception as e:
        print(e)
        error(logger_name, e)
    finally:
        prev_time, prstr = time_diff(start_time)
        info(logger_name, "Total time taken by the script: " + prstr)
        
def process_rids(rows, output_path, logger_name):
    auth_token = get_auth_token(logger_name)
    count = 0
    row_count_to_write = 0
    thread_list = []
    cur_row_count = 3
    wb = load_workbook(output_path)
    start_time = get_time_in_sec()
    for row in rows:
        rid = row["rid"]
        count = count + 1
        row_count_to_write = row_count_to_write + 1
        thread_list.append(ParentThread(auth_token, rid, logger_name))
        if (count == conf.thread_count):
            process_thread(thread_list)
            prev_time, prstr = time_diff(start_time)
            info(logger_name, f"Time taken to complete {count} packets requests : " + prstr)            
            count = 0
            thread_list = []
            start_time = get_time_in_sec()
        if (row_count_to_write == conf.row_count_to_write):
            row_count_to_write = 0
            cur_row_count = write_into_xls(output_path, result_list, cur_row_count, wb, logger_name)
            result_list.clear()
    if (len(thread_list) > 0):
        process_thread(thread_list)
        cur_row_count = write_into_xls(output_path, result_list, cur_row_count, wb, logger_name)
        result_list.clear() 
    if (len(result_list) > 0):
        cur_row_count = write_into_xls(output_path, result_list, cur_row_count, wb, logger_name)
        result_list.clear()               

def process_thread(thread_list):
    for thread in thread_list:
        thread.start()
    for thread in thread_list:
        thread.join()


def write_into_xls(output_path, r_list, cur_row_count, wb, logger_name):
    s_time = get_time_in_sec()
    ws = wb['Result']
    for result in r_list:
        ws.cell(cur_row_count, 1).value = result["rid"]
        ws.cell(cur_row_count, 2).value = result["status"]
        ws.cell(cur_row_count, 3).value = result["bio_status"] if result["bio_status"] else ""
        cur_row_count = cur_row_count + 1
    wb.save(output_path)
    prev_time, prstr = time_diff(s_time)
    info(logger_name, f"Time taken to complete {len(r_list)} rows to write : " + prstr)     
    return cur_row_count
 

def process_rid(rid, auth_token, logger_name):
    t1 = IdRepoThread(auth_token, rid, logger_name)
    t2 = PacketManagerThread(auth_token, rid, logger_name)
    t3 = PacketManagerBiometricsThread(auth_token, rid, logger_name)
    s_time = get_time_in_sec()
    t1.start()
    t2.start()
    t3.start()
    t1.join()
    t2.join()
    t3.join()
    prev_time, prstr = time_diff(s_time)
    debug(logger_name, f"Time taken to complete API calls : " + prstr)      
    idrepo_info = t1.value
    pkt_info = t2.value
    pkt_biometrics = t3.value
    s_time = get_time_in_sec()
    compare(idrepo_info, pkt_info, pkt_biometrics, rid, logger_name)
    prev_time, prstr = time_diff(s_time)
    debug(logger_name, f"Time taken to complete comparision : " + prstr)  

def compare(idrepo_info, pkt_info, pkt_biometrics, rid, logger_name):
    return_value = {
        "rid": rid,
        "status": "",
        "bio_status": ""
    }
    if (idrepo_info and pkt_info):
        return_value = compare_demoinfo(idrepo_info, pkt_info, rid, return_value, logger_name)
        return_value = compare_bioinfo(idrepo_info, pkt_biometrics, rid, return_value, logger_name)
    else:
        status = ""
        if (idrepo_info == None):
            info(logger_name, f"RID {rid} - IdRepo not found")
            status = "IdRepo not found"
        if (pkt_info == None):
            info(logger_name, f"RID {rid} - Packet Info not found")
            status = status + "\n" + "Packet Info not found" if (len(status) > 0) else "Packet Info not found"
        if (len(status) > 0):
            return_value["status"] = status
    result_list.append(return_value)       

def compare_bioinfo(idrepo_info, pkt_biometrics, rid, return_value, logger_name):
    try:
        pkt_segments_list = pkt_biometrics["segments"]
        idrepo_biometrics_value = idrepo_info["documents"][0]["value"]
        idrepo_decoded_biometrics = base64.urlsafe_b64decode(idrepo_biometrics_value + '=' * (-len(idrepo_biometrics_value) % 4))
        idrepo_biometrics = urllib.parse.unquote(idrepo_decoded_biometrics.decode())
        idrepo_bio_dict = xmltodict.parse(idrepo_biometrics)
        if (idrepo_bio_dict and "BIR" in idrepo_bio_dict and "BIR" in idrepo_bio_dict["BIR"]):
            idrepo_segments_list = idrepo_bio_dict["BIR"]["BIR"]
            is_matching = True
            for idrepo_segments in idrepo_segments_list:
                if isinstance(idrepo_segments, str):
                    idrepo_segments = xmltodict.parse(idrepo_biometrics)
                    idrepo_segments = idrepo_segments["BIR"]["BIR"]
                idrepo_bdbinfo = idrepo_segments["BDBInfo"]
                idrepo_creationdate = idrepo_bdbinfo["CreationDate"]
                idrepo_qualityscore = idrepo_bdbinfo["Quality"]["Score"]
                idrepo_modtype = idrepo_bdbinfo["Type"]
                idrepo_subtype = idrepo_bdbinfo["Subtype"]
                pkt_creation_date, pkt_quality_score = get_pkt_info_for(pkt_segments_list, idrepo_modtype, idrepo_subtype)
                if (pkt_creation_date and idrepo_creationdate and pkt_quality_score and idrepo_qualityscore):
                    idrepo_creationdate = idrepo_creationdate[0:idrepo_creationdate.rindex(".")]
                    pkt_creation_date = pkt_creation_date[0:pkt_creation_date.rindex(".")]
                    if (pkt_creation_date != idrepo_creationdate or str(pkt_quality_score) != idrepo_qualityscore):
                        is_matching = False

            if (is_matching):
                return_value["bio_status"] = "Matching"
                info(logger_name, f"RID {rid} - Biographic - IdRepo and PacketInfo are matching")
            else:
                return_value["bio_status"] = "Not Matching"
                info(logger_name, f"RID {rid} - Biographic - IdRepo and PacketInfo are not matching")
        else:
            return_value["bio_status"] = "Invalid IdRepo Data"
    except Exception as e:
        print(e, rid)
        error(logger_name, e)            
    return return_value

def get_pkt_info_for(pkt_segments_list, type, subtype):
    pkt_creation_date = ""
    pkt_quality_score = ""
    for pkt_segments in pkt_segments_list:
        pkt_bdbinfo = pkt_segments["bdbInfo"]
        pkt_modtype = pkt_bdbinfo["type"][0]
        pkt_subtype = pkt_bdbinfo["subtype"]
        if len(pkt_bdbinfo["subtype"]) > 0:
            pkt_subtype = pkt_bdbinfo["subtype"][0] + " " + pkt_bdbinfo["subtype"][1] if len(pkt_bdbinfo["subtype"]) > 1 else pkt_bdbinfo["subtype"][0]
        else:
            pkt_subtype = None

        if (type.lower() == "face"):
            if (type.lower() == pkt_modtype.lower()):
                pkt_creation_date = pkt_bdbinfo["creationDate"]
                pkt_quality_score = pkt_bdbinfo["quality"]["score"]
                break                
        elif (type.lower() == pkt_modtype.lower() and subtype.lower() == pkt_subtype.lower()):
            pkt_creation_date = pkt_bdbinfo["creationDate"]
            pkt_quality_score = pkt_bdbinfo["quality"]["score"]
            break
    return pkt_creation_date,pkt_quality_score
      

def compare_demoinfo(idrepo_info, pkt_info, rid, return_value, logger_name):
    identity = idrepo_info["identity"]
    identity_firstname = get_value_for(identity["firstName"]) if ("firstName" in identity) else ""
    identity_lastname = get_value_for(identity["lastName"]) if ("lastName" in identity) else ""
    identity_middlename = get_value_for(identity["middleName"]) if ("middleName" in identity) else ""
    identity_gender = get_value_for(identity["gender"]) if ("gender" in identity) else ""
    identity_presentprovince = get_value_for(identity["presentProvince"]) if ("presentProvince" in identity) else ""
    
    pkt_firstname = ""
    if ("firstName" in pkt_info["fields"]):
        pkt_firstname = pkt_info["fields"]["firstName"]
        pkt_firstname = json.loads(pkt_firstname)
        pkt_firstname = get_value_for(pkt_firstname)
    
    pkt_lastname = ""
    if ("lastName" in pkt_info["fields"]):
        pkt_lastname = pkt_info["fields"]["lastName"]
        pkt_lastname = json.loads(pkt_lastname)
        pkt_lastname = get_value_for(pkt_lastname)
    
    pkt_middlename = ""
    if ("middleName" in pkt_info["fields"]):
        pkt_middlename = pkt_info["fields"]["middleName"]
        pkt_middlename = json.loads(pkt_middlename)
        pkt_middlename = get_value_for(pkt_middlename)
    
    pkt_gender = ""
    if ("gender" in pkt_info["fields"]):
        pkt_gender = pkt_info["fields"]["gender"]
        pkt_gender = json.loads(pkt_gender)
        pkt_gender = get_value_for(pkt_gender)
    
    pkt_presentprovince = ""
    if ("presentProvince" in pkt_info["fields"]):
        pkt_presentprovince = pkt_info["fields"]["presentProvince"]
        pkt_presentprovince = json.loads(pkt_presentprovince)
        pkt_presentprovince = get_value_for(pkt_presentprovince)

    identity_result = {
        "firstName": identity_firstname,
        "lastName": identity_lastname,
        "middleName": identity_middlename,
        "gender": identity_gender,
        "province": identity_presentprovince
    }
    pktinfo_result = {
        "firstName": pkt_firstname,
        "lastName": pkt_lastname,
        "middleName": pkt_middlename,
        "gender": pkt_gender,
        "province": pkt_presentprovince
    }    
    # return_value["identity_result"] = identity_result
    # return_value["pktinfo_result"] = pktinfo_result
    if (identity_firstname != pkt_firstname or
            identity_lastname != pkt_lastname or
            identity_middlename != pkt_middlename or
            identity_gender != pkt_gender or
            identity_presentprovince != pkt_presentprovince):
        return_value["status"] = "Not matching"
        # return_value["identity_result"] = identity_result
        # return_value["pktinfo_result"] = pktinfo_result            
        info(logger_name, f"RID {rid} - Demographic - IdRepo and PacketInfo are not matching")
    else:
        return_value["status"] = "Matching"
        info(logger_name, f"RID {rid} - Demographic - IdRepo and PacketInfo are matching")
    return return_value

def get_value_for(list):
    for d in list:
        if (d.get("language") == "eng"):
            return d.get("value")
    return None

def get_idrepo_identity_by_rid(token, rid, logger_name):
    try:
        s_time = get_time_in_sec()
        # info(f"Identity request api called for {rid}")
        type="bio"
        url = '%s/idrepository/v1/identity/idvid/%s' % (conf.identity_server_url, rid)
        params = {
            'idType': 'rid',
            'type': 'bio'
        }
        cookies = {'Authorization': token}
        r = requests.get(url, cookies=cookies, verify=conf.ssl_verify, params=params)
        resp = parse_response(r)
        prev_time, prstr = time_diff(s_time)
        debug(logger_name, f"Time taken to complete Identity API call : " + prstr)
        return resp
    except Exception as ex:
        error(logger_name, f"Failed to get identity for RID {rid} - {ex}")
        return None 

def get_info_from_packet(token, field_names_list, rid, reg_type, logger_name):
    try:
        s_time = get_time_in_sec()
        # info(f"Packetmanager searchfield api called for {rid}")
        url = '%s/commons/v1/packetmanager/searchFields' % conf.pkt_mgr_server_url
        cookies = {'Authorization': token}
        ts = get_timestamp()
        req = {
            "id": "packetmanager.searchfield",
            "metadata": {},
            "request": get_searchfield_request(field_names_list, rid, reg_type),
            "requesttime": ts,
            "version": "1.0"
        }
        r = requests.post(url, cookies=cookies, json=req, verify=conf.ssl_verify)
        resp = parse_response(r)
        prev_time, prstr = time_diff(s_time)
        debug(logger_name, f"Time taken to complete SearchFields API call : " + prstr)        
        return resp
    except Exception as ex:
        error(logger_name, f"Failed to get packet info for RID {rid} - {ex}")
        return None

def get_biometrics_from_packet(token, rid, reg_type, logger_name):
    try:
        s_time = get_time_in_sec()
        # info(f"Packetmanager biometrics api called for {rid}")
        url = '%s/commons/v1/packetmanager/biometrics' % conf.pkt_mgr_server_url
        cookies = {'Authorization': token}
        ts = get_timestamp()
        req = {
            "id": "packetmanager.biometrics",
            "metadata": {},
            "request": get_biometrics_request(rid, reg_type),
            "requesttime": ts,
            "version": "1.0"
        }
        r = requests.post(url, cookies=cookies, json=req, verify=conf.ssl_verify)
        resp = parse_response(r)
        prev_time, prstr = time_diff(s_time)
        debug(logger_name, f"Time taken to complete Biometrics API call : " + prstr)        
        return resp
    except Exception as ex:
        error(logger_name, f"Failed to get packet biometrics for RID {rid} - {ex}")
        return None

def get_searchfield_request(field_names_list, rid, reg_type):
    return {
        "bypassCache": False,
        "fields": field_names_list,
        "id": rid,
        "process": reg_type,
        "source": "REGISTRATION_CLIENT"
    }

def get_biometrics_request(rid, reg_type):
    return {
        "bypassCache": False,
        "id": rid,
        "person": "individualBiometrics",
        "process": reg_type,
        "source": "REGISTRATION_CLIENT"
    }

def fetch_rids(from_date, to_date):
    try:
        query = f"""
                select reg_id as RID from idrepo.uin where cr_dtimes > '{from_date}' and cr_dtimes < '{to_date}'
                """
        reg_id_list = fetch_all("mosip_idrepo", query)
        return reg_id_list
    except Exception as ex:
        raise Exception(f"while fetching fetch_rids info... {ex}")    

if __name__ == "__main__":
    main()


